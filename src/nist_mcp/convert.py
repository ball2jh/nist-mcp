"""Multi-format document-to-Markdown converter.

Handles PDF (via pymupdf4llm), XLSX/XLS (openpyxl), CSV, JSON, XML, and
plain text.  Converted results are cached as ``.md`` files alongside the
source so repeated reads are cheap.
"""

from __future__ import annotations

import csv
import io
import json
import logging
import tempfile
import xml.dom.minidom
from pathlib import Path
from urllib.parse import urljoin

import httpx

from nist_mcp.safety import (
    ALLOWED_DOCUMENT_HOSTS,
    MAX_DOCUMENT_BYTES,
    validate_https_url,
    validate_page_range,
)

log = logging.getLogger(__name__)


# ---------------------------------------------------------------------------
# Download helper
# ---------------------------------------------------------------------------


async def download_file(url: str, dest: Path) -> Path:
    """Download *url* to *dest*.  Skip if *dest* already exists.

    Creates parent directories as needed.  Returns *dest*.
    """
    url = validate_https_url(url, allowed_hosts=ALLOWED_DOCUMENT_HOSTS)

    if dest.exists():
        log.debug("File already cached: %s", dest)
        return dest

    dest.parent.mkdir(parents=True, exist_ok=True)
    log.info("Downloading %s -> %s", url, dest)

    redirects_remaining = 5
    tmp_path: Path | None = None

    try:
        async with httpx.AsyncClient(follow_redirects=False, timeout=120) as client:
            while True:
                validate_https_url(url, allowed_hosts=ALLOWED_DOCUMENT_HOSTS)
                async with client.stream("GET", url) as resp:
                    if resp.is_redirect:
                        location = resp.headers.get("location")
                        if not location:
                            raise httpx.HTTPStatusError(
                                "Redirect response missing Location header",
                                request=resp.request,
                                response=resp,
                            )
                        redirects_remaining -= 1
                        if redirects_remaining < 0:
                            raise httpx.TooManyRedirects("Too many redirects")
                        url = urljoin(str(resp.url), location)
                        validate_https_url(url, allowed_hosts=ALLOWED_DOCUMENT_HOSTS)
                        continue

                    resp.raise_for_status()

                    with tempfile.NamedTemporaryFile(
                        "wb",
                        delete=False,
                        dir=str(dest.parent),
                        prefix=f".{dest.name}.",
                    ) as tmp:
                        tmp_path = Path(tmp.name)
                        total = 0
                        async for chunk in resp.aiter_bytes():
                            total += len(chunk)
                            if total > MAX_DOCUMENT_BYTES:
                                raise ValueError(
                                    "Downloaded document exceeds "
                                    f"{MAX_DOCUMENT_BYTES // (1024 * 1024)} MB limit."
                                )
                            tmp.write(chunk)
                    break
    except Exception:
        if tmp_path is not None:
            tmp_path.unlink(missing_ok=True)
        raise

    if tmp_path is None:
        raise RuntimeError("Download did not produce a file.")

    tmp_path.replace(dest)
    return dest


# ---------------------------------------------------------------------------
# Public conversion API
# ---------------------------------------------------------------------------


def convert_to_markdown(source_path: Path, pages: str | None = None) -> str:
    """Convert a file to Markdown based on its extension.

    Supported formats:

    - ``.pdf`` — pymupdf4llm (primary) with optional page range
    - ``.xlsx`` / ``.xls`` — openpyxl, renders sheets as markdown tables
    - ``.csv`` — stdlib csv, rendered as a markdown table
    - ``.json`` — pretty-printed JSON
    - ``.xml`` — pretty-printed XML
    - other — raw text content

    Args:
        source_path: Path to the source file.
        pages: Optional page range for PDFs (e.g. ``"1-50"``).

    Returns:
        Markdown string.
    """
    suffix = source_path.suffix.lower()

    # Check for a cached .md conversion (not for partial reads).
    if pages is None:
        cached = _cached_md_path(source_path)
        if cached.exists() and cached.stat().st_mtime >= source_path.stat().st_mtime:
            return cached.read_text(encoding="utf-8")

    if suffix == ".pdf":
        md = _convert_pdf(source_path, pages)
    elif suffix in (".xlsx", ".xls"):
        md = _convert_xlsx(source_path)
    elif suffix == ".csv":
        md = _convert_csv(source_path)
    elif suffix == ".json":
        md = _convert_json(source_path)
    elif suffix == ".xml":
        md = _convert_xml(source_path)
    else:
        md = source_path.read_text(encoding="utf-8", errors="replace")

    # Cache full conversions.
    if pages is None:
        cached = _cached_md_path(source_path)
        cached.write_text(md, encoding="utf-8")

    return md


def get_pdf_toc(pdf_path: Path) -> str:
    """Extract table of contents from a PDF.

    Returns a formatted Markdown outline.  If the PDF has no outline
    (bookmarks), returns a short notice.
    """
    import pymupdf

    doc = pymupdf.open(str(pdf_path))
    toc = doc.get_toc()  # list of [level, title, page]
    doc.close()

    if not toc:
        return "_This PDF has no table of contents (bookmarks)._"

    lines = ["# Table of Contents\n"]
    for level, title, page in toc:
        indent = "  " * (level - 1)
        lines.append(f"{indent}- {title} (p. {page})")

    return "\n".join(lines)


def get_pdf_section(pdf_path: Path, section_name: str) -> str:
    """Extract a specific section from a PDF by heading match.

    Finds the TOC entry whose title best matches *section_name*, determines
    the page range (from its page to the next same-or-higher-level heading),
    and extracts those pages.
    """
    import pymupdf

    doc = pymupdf.open(str(pdf_path))
    toc = doc.get_toc()
    total_pages = doc.page_count
    doc.close()

    if not toc:
        return (
            "_This PDF has no table of contents.  "
            "Use the `pages` parameter to request a specific page range._"
        )

    # Find the best matching TOC entry (case-insensitive substring match).
    needle = section_name.lower()
    match_idx: int | None = None
    for i, (_, title, _) in enumerate(toc):
        if needle in title.lower():
            match_idx = i
            break

    if match_idx is None:
        # Fall back to partial word match.
        for i, (_, title, _) in enumerate(toc):
            if any(w in title.lower() for w in needle.split()):
                match_idx = i
                break

    if match_idx is None:
        available = "\n".join(f"- {title}" for _, title, _ in toc[:30])
        return (
            f"Section **{section_name}** not found in the table of contents.\n\n"
            f"Available sections (first 30):\n{available}"
        )

    match_level, match_title, match_page = toc[match_idx]

    # Determine end page: next heading at the same or higher level.
    end_page = total_pages
    for j in range(match_idx + 1, len(toc)):
        next_level, _, next_page = toc[j]
        if next_level <= match_level:
            end_page = next_page  # exclusive — this heading starts a new section
            break

    # pymupdf4llm pages are 0-based.
    page_list = list(range(match_page - 1, end_page - 1)) or [match_page - 1]

    import pymupdf4llm

    md = pymupdf4llm.to_markdown(str(pdf_path), pages=page_list)
    return f"## {match_title}\n\n{md}"


# ---------------------------------------------------------------------------
# Private converters
# ---------------------------------------------------------------------------


def _convert_pdf(source_path: Path, pages: str | None) -> str:
    """Convert a PDF to Markdown via pymupdf4llm."""
    import pymupdf4llm

    kwargs: dict = {}
    if pages is not None:
        kwargs["pages"] = _parse_page_range(pages)

    return pymupdf4llm.to_markdown(str(source_path), **kwargs)


def _convert_xlsx(source_path: Path) -> str:
    """Convert an XLSX/XLS workbook to Markdown tables."""
    import openpyxl

    wb = openpyxl.load_workbook(str(source_path), read_only=True, data_only=True)
    parts: list[str] = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue

        parts.append(f"## {sheet_name}\n")

        # First row as headers.
        headers = [str(c) if c is not None else "" for c in rows[0]]
        parts.append("| " + " | ".join(headers) + " |")
        parts.append("| " + " | ".join("---" for _ in headers) + " |")

        for row in rows[1:]:
            cells = [str(c) if c is not None else "" for c in row]
            parts.append("| " + " | ".join(cells) + " |")

        parts.append("")  # blank line between sheets

    wb.close()
    return "\n".join(parts)


def _convert_csv(source_path: Path) -> str:
    """Convert a CSV file to a Markdown table."""
    text = source_path.read_text(encoding="utf-8", errors="replace")
    reader = csv.reader(io.StringIO(text))
    rows = list(reader)
    if not rows:
        return "_Empty CSV file._"

    headers = rows[0]
    lines = [
        "| " + " | ".join(headers) + " |",
        "| " + " | ".join("---" for _ in headers) + " |",
    ]
    for row in rows[1:]:
        # Pad short rows.
        cells = row + [""] * (len(headers) - len(row))
        lines.append("| " + " | ".join(cells[: len(headers)]) + " |")

    return "\n".join(lines)


def _convert_json(source_path: Path) -> str:
    """Pretty-print a JSON file."""
    raw = source_path.read_text(encoding="utf-8", errors="replace")
    try:
        data = json.loads(raw)
        formatted = json.dumps(data, indent=2, ensure_ascii=False)
    except json.JSONDecodeError:
        formatted = raw
    return f"```json\n{formatted}\n```"


def _convert_xml(source_path: Path) -> str:
    """Pretty-print an XML file."""
    raw = source_path.read_text(encoding="utf-8", errors="replace")
    try:
        dom = xml.dom.minidom.parseString(raw)
        formatted = dom.toprettyxml(indent="  ")
    except Exception:
        formatted = raw
    return f"```xml\n{formatted}\n```"


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------


def _cached_md_path(source_path: Path) -> Path:
    """Return the path for a cached .md conversion alongside *source_path*."""
    return source_path.with_suffix(source_path.suffix + ".md")


def _parse_page_range(pages: str) -> list[int]:
    """Parse a page range string into a list of 0-based page numbers.

    Supports formats like ``"1-50"``, ``"3"``, ``"1-5,10-12"``.
    Input is 1-based; output is 0-based.
    """
    return validate_page_range(pages)
