"""Scraper for NIST Cybersecurity Framework (CSF) 2.0 data.

Downloads the official CSF 2.0 XLSX export from NIST CSRC and parses it into
a three-level hierarchy: Function -> Category -> Subcategory.  Falls back to
a hardcoded dataset when the download is unavailable or the format is
unexpected.
"""

from __future__ import annotations

import io
import logging
import re
import sqlite3
from typing import Any

import httpx

log = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# Schema
# ---------------------------------------------------------------------------

CREATE_TABLE_SQL = """
CREATE TABLE IF NOT EXISTS csf (
    id            TEXT PRIMARY KEY,
    function_id   TEXT,
    function_name TEXT,
    category_id   TEXT,
    category_name TEXT,
    title         TEXT NOT NULL,
    level         TEXT NOT NULL   -- 'function', 'category', or 'subcategory'
);
CREATE INDEX IF NOT EXISTS idx_csf_function ON csf(function_id);
CREATE INDEX IF NOT EXISTS idx_csf_level ON csf(level);
"""

# ---------------------------------------------------------------------------
# Data source URL
# ---------------------------------------------------------------------------

_CSF_URL = "https://csrc.nist.gov/extensions/nudp/services/json/csf/download?olirids=all"
_TIMEOUT = 60

# ---------------------------------------------------------------------------
# Hardcoded CSF 2.0 reference data (fallback)
# ---------------------------------------------------------------------------

# The 6 CSF 2.0 Functions
_FUNCTIONS: list[tuple[str, str]] = [
    ("GV", "Govern"),
    ("ID", "Identify"),
    ("PR", "Protect"),
    ("DE", "Detect"),
    ("RS", "Respond"),
    ("RC", "Recover"),
]

# Categories: (category_id, function_id, category_name)
_CATEGORIES: list[tuple[str, str, str]] = [
    # Govern
    ("GV.OC", "GV", "Organizational Context"),
    ("GV.RM", "GV", "Risk Management Strategy"),
    ("GV.RR", "GV", "Roles, Responsibilities, and Authorities"),
    ("GV.PO", "GV", "Policy"),
    ("GV.OV", "GV", "Oversight"),
    ("GV.SC", "GV", "Cybersecurity Supply Chain Risk Management"),
    # Identify
    ("ID.AM", "ID", "Asset Management"),
    ("ID.RA", "ID", "Risk Assessment"),
    ("ID.IM", "ID", "Improvement"),
    # Protect
    ("PR.AA", "PR", "Identity Management, Authentication, and Access Control"),
    ("PR.AT", "PR", "Awareness and Training"),
    ("PR.DS", "PR", "Data Security"),
    ("PR.PS", "PR", "Platform Security"),
    ("PR.IR", "PR", "Technology Infrastructure Resilience"),
    # Detect
    ("DE.CM", "DE", "Continuous Monitoring"),
    ("DE.AE", "DE", "Adverse Event Analysis"),
    # Respond
    ("RS.MA", "RS", "Incident Management"),
    ("RS.AN", "RS", "Incident Analysis"),
    ("RS.CO", "RS", "Incident Response Reporting and Communication"),
    ("RS.MI", "RS", "Incident Mitigation"),
    # Recover
    ("RC.RP", "RC", "Incident Recovery Plan Execution"),
    ("RC.CO", "RC", "Incident Recovery Communication"),
]

# Subcategories: (subcategory_id, title)
# Complete CSF 2.0 subcategory list
_SUBCATEGORIES: list[tuple[str, str]] = [
    # GV.OC
    ("GV.OC-01", "The organizational mission is understood and informs cybersecurity risk management"),
    ("GV.OC-02", "Internal and external stakeholders are understood, and their needs and expectations regarding cybersecurity risk management are understood and considered"),
    ("GV.OC-03", "Legal, regulatory, and contractual requirements regarding cybersecurity - including privacy and civil liberties obligations - are understood and managed"),
    ("GV.OC-04", "Critical objectives, capabilities, and services that external stakeholders depend on or expect from the organization are understood and communicated"),
    ("GV.OC-05", "Outcomes, capabilities, and services that the organization depends on are understood and communicated"),
    # GV.RM
    ("GV.RM-01", "Risk management objectives are established and agreed to by organizational stakeholders"),
    ("GV.RM-02", "Risk appetite and risk tolerance statements are established, communicated, and maintained"),
    ("GV.RM-03", "Cybersecurity risk management activities and outcomes are included in enterprise risk management processes"),
    ("GV.RM-04", "Strategic direction that describes appropriate risk response options is established and communicated"),
    ("GV.RM-05", "Lines of communication across the organization are established for cybersecurity risks, including risks from suppliers and other third parties"),
    ("GV.RM-06", "A standardized method for calculating, documenting, categorizing, and prioritizing cybersecurity risks is established and communicated"),
    ("GV.RM-07", "Strategic opportunities (i.e., positive risks) are characterized and are included in organizational cybersecurity risk discussions"),
    # GV.RR
    ("GV.RR-01", "Organizational leadership is responsible and accountable for cybersecurity risk and fosters a culture that is risk-aware, ethical, and continually improving"),
    ("GV.RR-02", "Roles, responsibilities, and authorities related to cybersecurity risk management are established, communicated, understood, and enforced"),
    ("GV.RR-03", "Adequate resources are allocated commensurate with the cybersecurity risk strategy, roles and responsibilities, and policies"),
    ("GV.RR-04", "Cybersecurity is included in human resources practices"),
    # GV.PO
    ("GV.PO-01", "Policy for managing cybersecurity risks is established based on organizational context, cybersecurity strategy, and priorities and is communicated and enforced"),
    ("GV.PO-02", "Policy for managing cybersecurity risks is reviewed, updated, communicated, and enforced to reflect changes in requirements, threats, technology, and organizational mission"),
    # GV.OV
    ("GV.OV-01", "Cybersecurity risk management strategy outcomes are reviewed to inform and adjust strategy and direction"),
    ("GV.OV-02", "The cybersecurity risk management strategy is reviewed and adjusted to ensure coverage of organizational requirements and risks"),
    ("GV.OV-03", "Organizational cybersecurity risk management performance is evaluated and reviewed for adjustments needed"),
    # GV.SC
    ("GV.SC-01", "A cybersecurity supply chain risk management program, strategy, objectives, policies, and processes are established and agreed to by organizational stakeholders"),
    ("GV.SC-02", "Cybersecurity roles and responsibilities for suppliers, customers, and partners are established, communicated, and coordinated internally and externally"),
    ("GV.SC-03", "Cybersecurity supply chain risk management is integrated into cybersecurity and enterprise risk management, risk assessment, and improvement processes"),
    ("GV.SC-04", "Suppliers are known and prioritized by criticality"),
    ("GV.SC-05", "Requirements to address cybersecurity risks in supply chains are established, prioritized, and integrated into contracts and other types of agreements with suppliers and other relevant third parties"),
    ("GV.SC-06", "Planning and due diligence are performed to reduce risks before entering into formal supplier or other third-party relationships"),
    ("GV.SC-07", "The risks posed by a supplier, their products and services, and other third parties are understood, recorded, prioritized, assessed, responded to, and monitored over the course of the relationship"),
    ("GV.SC-08", "Relevant suppliers and other third parties are included in incident planning, response, and recovery activities"),
    ("GV.SC-09", "Supply chain security practices are integrated into cybersecurity and enterprise risk management programs, and their performance is monitored throughout the technology product and service life cycle"),
    ("GV.SC-10", "Cybersecurity supply chain risk management plans include provisions for activities that occur after the conclusion of a partnership or service agreement"),
    # ID.AM
    ("ID.AM-01", "Inventories of hardware managed by the organization are maintained"),
    ("ID.AM-02", "Inventories of software, services, and systems managed by the organization are maintained"),
    ("ID.AM-03", "Representations of the organization's authorized network communication and internal and external network data flows are maintained"),
    ("ID.AM-04", "Inventories of services provided by suppliers are maintained"),
    ("ID.AM-05", "Assets are prioritized based on classification, criticality, resources, and impact on the mission"),
    ("ID.AM-07", "Inventories of data and corresponding metadata for designated data types are maintained"),
    ("ID.AM-08", "Systems, hardware, software, services, and data are managed throughout their life cycles"),
    # ID.RA
    ("ID.RA-01", "Vulnerabilities in assets are identified, validated, and recorded"),
    ("ID.RA-02", "Cyber threat intelligence is received from information sharing forums and sources"),
    ("ID.RA-03", "Internal and external threats to the organization are identified and recorded"),
    ("ID.RA-04", "Potential impacts and likelihoods of threats exploiting vulnerabilities are identified and recorded"),
    ("ID.RA-05", "Threats, vulnerabilities, likelihoods, and impacts are used to understand inherent risk and inform risk response prioritization"),
    ("ID.RA-06", "Risk responses are chosen, prioritized, planned, tracked, and communicated"),
    ("ID.RA-07", "Changes and exceptions are managed, assessed for risk impact, recorded, and tracked"),
    ("ID.RA-08", "Processes for receiving, analyzing, and responding to vulnerability disclosures are established"),
    ("ID.RA-09", "The authenticity and integrity of hardware and software are assessed prior to acquisition and use"),
    ("ID.RA-10", "Critical suppliers are assessed prior to acquisition"),
    # ID.IM
    ("ID.IM-01", "Improvements are identified from evaluations"),
    ("ID.IM-02", "Improvements are identified from security tests and exercises, including those done in coordination with suppliers and relevant third parties"),
    ("ID.IM-03", "Improvements are identified from execution of operational processes, procedures, and activities"),
    ("ID.IM-04", "Incident response plans and other cybersecurity plans that affect operations are established, communicated, maintained, and improved"),
    # PR.AA
    ("PR.AA-01", "Identities and credentials for authorized users, services, and hardware are managed by the organization"),
    ("PR.AA-02", "Identities are proofed and bound to credentials based on the context of interactions"),
    ("PR.AA-03", "Users, services, and hardware are authenticated"),
    ("PR.AA-04", "Identity assertions are protected, conveyed, and verified"),
    ("PR.AA-05", "Access permissions, entitlements, and authorizations are defined in a policy, managed, enforced, and reviewed, and incorporate the principles of least privilege and separation of duties"),
    ("PR.AA-06", "Physical access to assets is managed, monitored, and enforced commensurate with risk"),
    # PR.AT
    ("PR.AT-01", "Personnel are provided with awareness and training so that they possess the knowledge and skills to perform general tasks with cybersecurity risks in mind"),
    ("PR.AT-02", "Individuals in specialized roles are provided with awareness and training so that they possess the knowledge and skills to perform relevant tasks with cybersecurity risks in mind"),
    # PR.DS
    ("PR.DS-01", "The confidentiality, integrity, and availability of data-at-rest are protected"),
    ("PR.DS-02", "The confidentiality, integrity, and availability of data-in-transit are protected"),
    ("PR.DS-10", "The confidentiality, integrity, and availability of data-in-use are protected"),
    ("PR.DS-11", "Backups of data are created, protected, maintained, and tested"),
    # PR.PS
    ("PR.PS-01", "Configuration management practices are established and applied"),
    ("PR.PS-02", "Software is maintained, replaced, and removed commensurate with risk"),
    ("PR.PS-03", "Hardware is maintained, replaced, and removed commensurate with risk"),
    ("PR.PS-04", "Log records are generated and made available for continuous monitoring"),
    ("PR.PS-05", "Installation and execution of unauthorized software are prevented"),
    ("PR.PS-06", "Secure software development practices are integrated, and their performance is monitored throughout the software development life cycle"),
    # PR.IR
    ("PR.IR-01", "Networks and environments are protected from unauthorized logical access and usage"),
    ("PR.IR-02", "The organization's technology assets are protected from environmental threats"),
    ("PR.IR-03", "Mechanisms are implemented to achieve resilience requirements in normal and adverse situations"),
    ("PR.IR-04", "Adequate resource capacity to ensure availability is maintained"),
    # DE.CM
    ("DE.CM-01", "Networks and network services are monitored to find potentially adverse events"),
    ("DE.CM-02", "The physical environment is monitored to find potentially adverse events"),
    ("DE.CM-03", "Personnel activity and technology usage are monitored to find potentially adverse events"),
    ("DE.CM-06", "External service provider activities and services are monitored to find potentially adverse events"),
    ("DE.CM-09", "Computing hardware and software, runtime environments, and their data are monitored to find potentially adverse events"),
    # DE.AE
    ("DE.AE-02", "Potentially adverse events are analyzed to better understand associated activities"),
    ("DE.AE-03", "Information is correlated from multiple sources"),
    ("DE.AE-04", "The estimated impact and scope of adverse events are understood"),
    ("DE.AE-06", "Information on adverse events is provided to authorized staff and tools"),
    ("DE.AE-07", "Cyber threat intelligence and other contextual information are integrated into the analysis"),
    ("DE.AE-08", "Incidents are declared when adverse events meet the defined incident criteria"),
    # RS.MA
    ("RS.MA-01", "The incident response plan is executed in coordination with relevant third parties once an incident is declared or detected"),
    ("RS.MA-02", "Incident reports are triaged and validated"),
    ("RS.MA-03", "Incidents are categorized and prioritized"),
    ("RS.MA-04", "Incidents are escalated or elevated as needed"),
    ("RS.MA-05", "The criteria for initiating incident recovery are applied"),
    # RS.AN
    ("RS.AN-03", "Analysis is performed to establish what has taken place during an incident and the root cause of the incident"),
    ("RS.AN-06", "Actions performed during an investigation are recorded, and the records' integrity and provenance are preserved"),
    ("RS.AN-07", "Incident data and metadata are collected, and their integrity and provenance are preserved"),
    ("RS.AN-08", "An incident's magnitude is estimated and validated"),
    # RS.CO
    ("RS.CO-02", "Internal and external stakeholders are notified of incidents"),
    ("RS.CO-03", "Information is shared with designated internal and external stakeholders"),
    # RS.MI
    ("RS.MI-01", "Incidents are contained"),
    ("RS.MI-02", "Incidents are eradicated"),
    # RC.RP
    ("RC.RP-01", "The recovery portion of the incident response plan is executed once initiated from the incident response process"),
    ("RC.RP-02", "Recovery actions are selected, scoped, prioritized, and performed"),
    ("RC.RP-03", "The integrity of backups and other restoration assets is verified before using them for restoration"),
    ("RC.RP-04", "Critical mission functions and cybersecurity risk management are considered to establish post-incident operational norms"),
    ("RC.RP-05", "The integrity of restored assets is verified, systems and services are restored, and normal operating status is confirmed"),
    ("RC.RP-06", "The end of incident recovery is declared based on criteria, and incident-related documentation is completed"),
    # RC.CO
    ("RC.CO-03", "Recovery activities and progress in restoring operational capabilities are communicated to designated internal and external stakeholders"),
    ("RC.CO-04", "Public updates on incident recovery are shared using approved methods and messaging"),
]


# ---------------------------------------------------------------------------
# XLSX parser
# ---------------------------------------------------------------------------


def _parse_csf_xlsx(data: bytes) -> list[dict[str, Any]] | None:
    """Attempt to parse the NIST CSF XLSX and return row dicts.

    The NIST CSRC XLSX has columns: Function, Category, Subcategory, etc.
    IDs are embedded in the cell text:
      - Function: "GOVERN (GV): The organization's ..."
      - Category: "Organizational Context (GV.OC): The circumstances ..."
      - Subcategory: "GV.OC-01: The organizational mission ..."

    Returns *None* if the format is unexpected so the caller can fall back.
    """
    try:
        from openpyxl import load_workbook
    except ImportError:
        log.warning("openpyxl not installed; cannot parse XLSX")
        return None

    wb = load_workbook(filename=io.BytesIO(data), read_only=True, data_only=True)

    # Patterns for extracting IDs from cell text
    func_re = re.compile(r"\(([A-Z]{2})\)\s*:")  # "GOVERN (GV): ..."
    cat_re = re.compile(r"\(([A-Z]{2}\.[A-Z]{2})\)\s*:")  # "... (GV.OC): ..."
    subcat_re = re.compile(r"^([A-Z]{2}\.[A-Z]{2}-\d+)\s*:")  # "GV.OC-01: ..."

    rows_out: list[dict[str, Any]] = []

    # Find the data sheet (skip "Introduction")
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]

        # Identify column layout from the header row
        func_col: int | None = None
        cat_col: int | None = None
        subcat_col: int | None = None

        for row_idx, row in enumerate(ws.iter_rows(values_only=True)):
            str_row = [str(c).strip() if c else "" for c in row]

            # Look for the header row with "Function", "Category", "Subcategory"
            if any("function" in c.lower() for c in str_row) and any(
                "category" in c.lower() for c in str_row
            ):
                for i, cell in enumerate(str_row):
                    cl = cell.lower()
                    if cl == "function":
                        func_col = i
                    elif cl == "category":
                        cat_col = i
                    elif cl == "subcategory":
                        subcat_col = i
                continue

            if func_col is None:
                continue  # Haven't found header yet

            # Parse data rows
            func_cell = str_row[func_col] if func_col is not None and func_col < len(str_row) else ""
            cat_cell = str_row[cat_col] if cat_col is not None and cat_col < len(str_row) else ""
            subcat_cell = str_row[subcat_col] if subcat_col is not None and subcat_col < len(str_row) else ""

            # Extract function
            if func_cell:
                m = func_re.search(func_cell)
                if m:
                    fid = m.group(1)
                    # Title is everything before the parenthetical ID
                    title = func_cell[: m.start()].strip()
                    rows_out.append({"id": fid, "title": title, "_type": "function"})

            # Extract category
            if cat_cell:
                m = cat_re.search(cat_cell)
                if m:
                    cid = m.group(1)
                    title = cat_cell[: cat_cell.index("(")].strip()
                    rows_out.append({"id": cid, "title": title, "_type": "category"})

            # Extract subcategory
            if subcat_cell:
                m = subcat_re.match(subcat_cell)
                if m:
                    sid = m.group(1)
                    title = subcat_cell[m.end():].strip()
                    rows_out.append({"id": sid, "title": title, "_type": "subcategory"})

    wb.close()

    if not rows_out:
        return None
    return rows_out


def _build_from_hardcoded() -> list[dict[str, Any]]:
    """Build CSF rows from the hardcoded reference data."""
    func_map = {fid: fname for fid, fname in _FUNCTIONS}
    cat_map: dict[str, tuple[str, str]] = {}  # cat_id -> (func_id, cat_name)
    for cat_id, func_id, cat_name in _CATEGORIES:
        cat_map[cat_id] = (func_id, cat_name)

    rows: list[dict[str, Any]] = []

    # Functions
    for fid, fname in _FUNCTIONS:
        rows.append({
            "id": fid,
            "function_id": fid,
            "function_name": fname,
            "category_id": None,
            "category_name": None,
            "title": fname,
            "level": "function",
        })

    # Categories
    for cat_id, func_id, cat_name in _CATEGORIES:
        rows.append({
            "id": cat_id,
            "function_id": func_id,
            "function_name": func_map[func_id],
            "category_id": cat_id,
            "category_name": cat_name,
            "title": cat_name,
            "level": "category",
        })

    # Subcategories
    for sub_id, title in _SUBCATEGORIES:
        # Parse "GV.OC-01" -> function_id="GV", category_id="GV.OC"
        parts = sub_id.split("-", 1)
        cat_id = parts[0]  # e.g. "GV.OC"
        func_id = cat_id.split(".")[0]  # e.g. "GV"
        func_name = func_map.get(func_id, "")
        cat_info = cat_map.get(cat_id)
        cat_name = cat_info[1] if cat_info else ""

        rows.append({
            "id": sub_id,
            "function_id": func_id,
            "function_name": func_name,
            "category_id": cat_id,
            "category_name": cat_name,
            "title": title,
            "level": "subcategory",
        })

    return rows


def _build_from_xlsx(parsed: list[dict[str, Any]]) -> list[dict[str, Any]]:
    """Enrich XLSX-parsed rows with hierarchy information.

    Each parsed item has: id, title, _type (function/category/subcategory).
    """
    # Build lookup tables from parsed data and hardcoded fallbacks
    func_map = {fid: fname for fid, fname in _FUNCTIONS}
    cat_map: dict[str, tuple[str, str]] = {}
    for cat_id, func_id, cat_name in _CATEGORIES:
        cat_map[cat_id] = (func_id, cat_name)

    # First pass: learn names from parsed data
    parsed_func_names: dict[str, str] = {}
    parsed_cat_names: dict[str, str] = {}
    for item in parsed:
        t = item.get("_type", "")
        if t == "function":
            parsed_func_names[item["id"]] = item["title"]
        elif t == "category":
            parsed_cat_names[item["id"]] = item["title"]

    rows: list[dict[str, Any]] = []
    seen: set[str] = set()

    for item in parsed:
        csf_id: str = item["id"]
        title: str = item["title"]
        item_type: str = item.get("_type", "")

        if csf_id in seen:
            continue
        seen.add(csf_id)

        if item_type == "function":
            func_name = title or func_map.get(csf_id, "")
            rows.append({
                "id": csf_id,
                "function_id": csf_id,
                "function_name": func_name,
                "category_id": None,
                "category_name": None,
                "title": func_name,
                "level": "function",
            })

        elif item_type == "category":
            func_id = csf_id.split(".")[0]
            func_name = parsed_func_names.get(func_id, func_map.get(func_id, ""))
            cat_name = title or (cat_map.get(csf_id, ("", ""))[1])

            # Ensure function row exists
            if func_id not in seen:
                seen.add(func_id)
                rows.append({
                    "id": func_id,
                    "function_id": func_id,
                    "function_name": func_name,
                    "category_id": None,
                    "category_name": None,
                    "title": func_name,
                    "level": "function",
                })

            rows.append({
                "id": csf_id,
                "function_id": func_id,
                "function_name": func_name,
                "category_id": csf_id,
                "category_name": cat_name,
                "title": cat_name,
                "level": "category",
            })

        elif item_type == "subcategory":
            parts = csf_id.split("-", 1)
            cat_id = parts[0]
            func_id = cat_id.split(".")[0]
            func_name = parsed_func_names.get(func_id, func_map.get(func_id, ""))
            cat_name = parsed_cat_names.get(cat_id, "")
            if not cat_name:
                ci = cat_map.get(cat_id)
                cat_name = ci[1] if ci else ""

            # Ensure parent rows exist
            if func_id not in seen:
                seen.add(func_id)
                rows.append({
                    "id": func_id,
                    "function_id": func_id,
                    "function_name": func_name,
                    "category_id": None,
                    "category_name": None,
                    "title": func_name,
                    "level": "function",
                })
            if cat_id not in seen:
                seen.add(cat_id)
                rows.append({
                    "id": cat_id,
                    "function_id": func_id,
                    "function_name": func_name,
                    "category_id": cat_id,
                    "category_name": cat_name,
                    "title": cat_name,
                    "level": "category",
                })

            rows.append({
                "id": csf_id,
                "function_id": func_id,
                "function_name": func_name,
                "category_id": cat_id,
                "category_name": cat_name,
                "title": title,
                "level": "subcategory",
            })

    return rows


# ---------------------------------------------------------------------------
# Public API
# ---------------------------------------------------------------------------


def scrape_csf(db: sqlite3.Connection) -> int:
    """Download CSF XLSX and populate the ``csf`` table.

    Falls back to hardcoded CSF 2.0 data if the download fails.
    Returns the number of rows inserted.
    """
    rows: list[dict[str, Any]] | None = None

    # Try the live download first
    try:
        log.info("Downloading CSF 2.0 XLSX from NIST ...")
        client = httpx.Client(timeout=_TIMEOUT, follow_redirects=True)
        resp = client.get(_CSF_URL)
        resp.raise_for_status()
        client.close()

        parsed = _parse_csf_xlsx(resp.content)
        if parsed:
            rows = _build_from_xlsx(parsed)
            log.info("Parsed %d rows from CSF XLSX", len(rows))
        else:
            log.warning("XLSX parsed but no CSF IDs found; falling back to hardcoded data")
    except Exception:
        log.warning("CSF download failed; using hardcoded CSF 2.0 data", exc_info=True)

    if rows is None:
        rows = _build_from_hardcoded()
        log.info("Using hardcoded CSF 2.0 data: %d rows", len(rows))

    # Insert into DB
    db.execute("DELETE FROM csf")
    db.executemany(
        """
        INSERT INTO csf (id, function_id, function_name, category_id,
                         category_name, title, level)
        VALUES (:id, :function_id, :function_name, :category_id,
                :category_name, :title, :level)
        """,
        rows,
    )
    db.commit()
    log.info("Inserted %d CSF entries", len(rows))
    return len(rows)


# ---------------------------------------------------------------------------
# Standalone test harness
# ---------------------------------------------------------------------------

if __name__ == "__main__":
    logging.basicConfig(level=logging.INFO, format="%(levelname)s: %(message)s")
    conn = sqlite3.connect(":memory:")
    conn.executescript(CREATE_TABLE_SQL)
    n = scrape_csf(conn)
    print(f"Inserted {n} CSF entries")

    for level in ("function", "category", "subcategory"):
        cur = conn.execute("SELECT COUNT(*) FROM csf WHERE level = ?", (level,))
        print(f"  {level}: {cur.fetchone()[0]}")

    cur = conn.execute("SELECT id, title FROM csf WHERE level = 'subcategory' LIMIT 5")
    for row in cur:
        print(f"  {row}")
    conn.close()
