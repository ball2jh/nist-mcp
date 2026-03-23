"""Scraper for control mappings between NIST frameworks.

Downloads the official CSF / Privacy Framework to SP 800-53r5 crosswalk XLSX
from NIST CSRC and parses it into the ``mappings`` table.  Falls back to a
hardcoded subset of the most important relationships when the download is
unavailable.
"""

from __future__ import annotations

import io
import logging
import re
import sqlite3
from typing import Any

import httpx

log = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# Schema
# ---------------------------------------------------------------------------

CREATE_TABLE_SQL = """
CREATE TABLE IF NOT EXISTS mappings (
    id               INTEGER PRIMARY KEY AUTOINCREMENT,
    source_framework TEXT NOT NULL,
    source_id        TEXT NOT NULL,
    target_framework TEXT NOT NULL,
    target_id        TEXT NOT NULL,
    relationship     TEXT NOT NULL DEFAULT 'related'
);
CREATE INDEX IF NOT EXISTS idx_map_source ON mappings(source_framework, source_id);
CREATE INDEX IF NOT EXISTS idx_map_target ON mappings(target_framework, target_id);
"""

# ---------------------------------------------------------------------------
# Data source
# ---------------------------------------------------------------------------

_MAPPINGS_URL = (
    "https://csrc.nist.gov/files/pubs/sp/800/53/r5/upd1/final/docs/"
    "csf-pf-to-sp800-53r5-mappings.xlsx"
)
_TIMEOUT = 90

# ---------------------------------------------------------------------------
# XLSX parser
# ---------------------------------------------------------------------------

# Regex for subcategory-style IDs in cell text
# CSF v1.1: "ID.AM-1", "PR.AC-3"  /  CSF v2.0: "GV.OC-01", "PR.AA-05"
# PF v1.0: "ID.IM-P1", "GV.PO-P3"
_SUBCAT_RE = re.compile(r"[A-Z]{2}\.[A-Z]{2}-[A-Z]?\d+")
# SP 800-53 control IDs: "AC-1", "AC-2(3)", "CM-8(4)", "PM-5"
_CTRL_RE = re.compile(r"[A-Z]{2}-\d+(?:\(\d+\))?")


def _parse_mappings_xlsx(data: bytes) -> list[dict[str, str]] | None:
    """Parse the NIST crosswalk XLSX into mapping row dicts.

    The XLSX has sheets:
    - "CSF to SP 800-53r5": columns Function, Category, Subcategory, Control
    - "PF to SP 800-53r5": columns Function, (blank), Category, Subcategory, Control

    Returns *None* if format is unexpected.
    """
    try:
        from openpyxl import load_workbook
    except ImportError:
        log.warning("openpyxl not installed; cannot parse XLSX")
        return None

    wb = load_workbook(filename=io.BytesIO(data), read_only=True, data_only=True)
    rows: list[dict[str, str]] = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        sheet_lower = sheet_name.lower()

        # Determine source framework from sheet name
        if "csf" in sheet_lower:
            source_fw = "CSF.1.1"
        elif "pf" in sheet_lower or "priv" in sheet_lower:
            source_fw = "PF.1.0"
        else:
            continue  # Skip README etc.

        # Find column indices from the header rows.  Headers may be
        # split across rows 0-4.  We look for a cell whose *entire*
        # content is "Subcategory" (or starts with it) for the source
        # column, and for a cell containing "800-53" for the control
        # column.  Row 0 is typically a title spanning the whole sheet,
        # so skip cells that are very long (>80 chars) to avoid false
        # matches in title text.
        subcat_col: int | None = None
        ctrl_col: int | None = None
        data_start_row = 0

        all_rows: list[list[str]] = []
        for row in ws.iter_rows(values_only=True):
            str_row = [str(c).strip() if c else "" for c in row]
            all_rows.append(str_row)

        for hrow_idx, hrow in enumerate(all_rows[:5]):
            for i, cell in enumerate(hrow):
                cl = cell.lower()
                # Match "Subcategory" as a standalone column header
                if cl == "subcategory" and subcat_col is None:
                    subcat_col = i
                    data_start_row = max(data_start_row, hrow_idx + 1)
                # Match 800-53 control column — skip overly long cells
                # (those are title rows, not header cells)
                if "800-53" in cl and len(cell) < 80 and ctrl_col is None:
                    ctrl_col = i
                    data_start_row = max(data_start_row, hrow_idx + 1)

        if subcat_col is None or ctrl_col is None:
            log.debug("Sheet %s: could not identify columns (subcat=%s, ctrl=%s)",
                      sheet_name, subcat_col, ctrl_col)
            continue

        log.debug("Sheet %s: subcat_col=%d, ctrl_col=%d, data starts at row %d",
                   sheet_name, subcat_col, ctrl_col, data_start_row)

        # Process data rows (skip headers)
        for row_idx, str_row in enumerate(all_rows):
            if row_idx < data_start_row:
                continue

            # Get cell contents
            subcat_cell = str_row[subcat_col] if subcat_col < len(str_row) else ""
            ctrl_cell = str_row[ctrl_col] if ctrl_col < len(str_row) else ""

            # Extract subcategory ID from the beginning of the cell text
            # e.g., "ID.AM-1: Physical devices and systems..."
            subcat_match = _SUBCAT_RE.search(subcat_cell)
            if not subcat_match:
                continue

            source_id = subcat_match.group(0)

            # Extract control IDs from the control cell (comma-separated)
            # e.g., "CM-8, PM-5" or "AC-4, CA-3, CA-9, PL-8, SA-17"
            target_ids = _CTRL_RE.findall(ctrl_cell)
            if not target_ids:
                continue

            for tid in target_ids:
                ctrl_id = tid.lower()
                # Convert AC-2(3) -> ac-2.3 to match OSCAL ID format
                ctrl_id = ctrl_id.replace("(", ".").replace(")", "")
                rows.append({
                    "source_framework": source_fw,
                    "source_id": source_id,
                    "target_framework": "SP.800-53.r5",
                    "target_id": ctrl_id,
                    "relationship": "related",
                })

    wb.close()

    if not rows:
        return None

    # Deduplicate
    seen: set[tuple[str, ...]] = set()
    unique: list[dict[str, str]] = []
    for r in rows:
        key = (r["source_framework"], r["source_id"], r["target_id"])
        if key not in seen:
            seen.add(key)
            unique.append(r)

    return unique


# ---------------------------------------------------------------------------
# Fallback: hardcoded CSF 2.0 -> SP 800-53 mappings
# ---------------------------------------------------------------------------

_FALLBACK_MAPPINGS: list[tuple[str, str, str, str, str]] = [
    # (source_framework, source_id, target_framework, target_id, relationship)
    # Govern
    ("CSF.2.0", "GV.OC-01", "SP.800-53.r5", "pm-7", "related"),
    ("CSF.2.0", "GV.OC-02", "SP.800-53.r5", "pm-25", "related"),
    ("CSF.2.0", "GV.OC-03", "SP.800-53.r5", "pm-26", "related"),
    ("CSF.2.0", "GV.OC-04", "SP.800-53.r5", "pm-7", "related"),
    ("CSF.2.0", "GV.OC-05", "SP.800-53.r5", "pm-7", "related"),
    ("CSF.2.0", "GV.RM-01", "SP.800-53.r5", "pm-9", "related"),
    ("CSF.2.0", "GV.RM-02", "SP.800-53.r5", "pm-9", "related"),
    ("CSF.2.0", "GV.RM-03", "SP.800-53.r5", "pm-28", "related"),
    ("CSF.2.0", "GV.RR-01", "SP.800-53.r5", "pm-2", "related"),
    ("CSF.2.0", "GV.RR-02", "SP.800-53.r5", "pm-2", "related"),
    ("CSF.2.0", "GV.PO-01", "SP.800-53.r5", "pl-1", "related"),
    ("CSF.2.0", "GV.PO-02", "SP.800-53.r5", "pl-1", "related"),
    ("CSF.2.0", "GV.SC-01", "SP.800-53.r5", "sr-1", "related"),
    ("CSF.2.0", "GV.SC-02", "SP.800-53.r5", "sr-2", "related"),
    ("CSF.2.0", "GV.SC-03", "SP.800-53.r5", "sr-3", "related"),
    ("CSF.2.0", "GV.SC-04", "SP.800-53.r5", "sr-6", "related"),
    # Identify
    ("CSF.2.0", "ID.AM-01", "SP.800-53.r5", "cm-8", "related"),
    ("CSF.2.0", "ID.AM-02", "SP.800-53.r5", "cm-8", "related"),
    ("CSF.2.0", "ID.AM-03", "SP.800-53.r5", "ac-4", "related"),
    ("CSF.2.0", "ID.AM-04", "SP.800-53.r5", "pm-5", "related"),
    ("CSF.2.0", "ID.AM-05", "SP.800-53.r5", "ra-2", "related"),
    ("CSF.2.0", "ID.RA-01", "SP.800-53.r5", "ra-5", "related"),
    ("CSF.2.0", "ID.RA-02", "SP.800-53.r5", "pm-16", "related"),
    ("CSF.2.0", "ID.RA-03", "SP.800-53.r5", "ra-3", "related"),
    ("CSF.2.0", "ID.RA-04", "SP.800-53.r5", "ra-3", "related"),
    ("CSF.2.0", "ID.RA-05", "SP.800-53.r5", "ra-3", "related"),
    ("CSF.2.0", "ID.RA-06", "SP.800-53.r5", "pm-4", "related"),
    # Protect
    ("CSF.2.0", "PR.AA-01", "SP.800-53.r5", "ia-1", "related"),
    ("CSF.2.0", "PR.AA-02", "SP.800-53.r5", "ia-12", "related"),
    ("CSF.2.0", "PR.AA-03", "SP.800-53.r5", "ia-2", "related"),
    ("CSF.2.0", "PR.AA-04", "SP.800-53.r5", "ia-8", "related"),
    ("CSF.2.0", "PR.AA-05", "SP.800-53.r5", "ac-3", "related"),
    ("CSF.2.0", "PR.AA-06", "SP.800-53.r5", "pe-3", "related"),
    ("CSF.2.0", "PR.AT-01", "SP.800-53.r5", "at-2", "related"),
    ("CSF.2.0", "PR.AT-02", "SP.800-53.r5", "at-3", "related"),
    ("CSF.2.0", "PR.DS-01", "SP.800-53.r5", "sc-28", "related"),
    ("CSF.2.0", "PR.DS-02", "SP.800-53.r5", "sc-8", "related"),
    ("CSF.2.0", "PR.DS-10", "SP.800-53.r5", "sc-28", "related"),
    ("CSF.2.0", "PR.DS-11", "SP.800-53.r5", "cp-9", "related"),
    ("CSF.2.0", "PR.PS-01", "SP.800-53.r5", "cm-1", "related"),
    ("CSF.2.0", "PR.PS-02", "SP.800-53.r5", "si-2", "related"),
    ("CSF.2.0", "PR.PS-04", "SP.800-53.r5", "au-1", "related"),
    ("CSF.2.0", "PR.PS-05", "SP.800-53.r5", "cm-7", "related"),
    ("CSF.2.0", "PR.PS-06", "SP.800-53.r5", "sa-8", "related"),
    ("CSF.2.0", "PR.IR-01", "SP.800-53.r5", "ac-4", "related"),
    ("CSF.2.0", "PR.IR-02", "SP.800-53.r5", "pe-1", "related"),
    ("CSF.2.0", "PR.IR-03", "SP.800-53.r5", "cp-2", "related"),
    ("CSF.2.0", "PR.IR-04", "SP.800-53.r5", "cp-2", "related"),
    # Detect
    ("CSF.2.0", "DE.CM-01", "SP.800-53.r5", "si-4", "related"),
    ("CSF.2.0", "DE.CM-02", "SP.800-53.r5", "pe-6", "related"),
    ("CSF.2.0", "DE.CM-03", "SP.800-53.r5", "ac-2", "related"),
    ("CSF.2.0", "DE.CM-06", "SP.800-53.r5", "sa-9", "related"),
    ("CSF.2.0", "DE.CM-09", "SP.800-53.r5", "si-4", "related"),
    ("CSF.2.0", "DE.AE-02", "SP.800-53.r5", "ir-4", "related"),
    ("CSF.2.0", "DE.AE-03", "SP.800-53.r5", "au-6", "related"),
    ("CSF.2.0", "DE.AE-04", "SP.800-53.r5", "cp-2", "related"),
    ("CSF.2.0", "DE.AE-06", "SP.800-53.r5", "ir-6", "related"),
    ("CSF.2.0", "DE.AE-07", "SP.800-53.r5", "pm-16", "related"),
    ("CSF.2.0", "DE.AE-08", "SP.800-53.r5", "ir-5", "related"),
    # Respond
    ("CSF.2.0", "RS.MA-01", "SP.800-53.r5", "ir-4", "related"),
    ("CSF.2.0", "RS.MA-02", "SP.800-53.r5", "ir-5", "related"),
    ("CSF.2.0", "RS.MA-03", "SP.800-53.r5", "ir-5", "related"),
    ("CSF.2.0", "RS.AN-03", "SP.800-53.r5", "ir-4", "related"),
    ("CSF.2.0", "RS.CO-02", "SP.800-53.r5", "ir-6", "related"),
    ("CSF.2.0", "RS.CO-03", "SP.800-53.r5", "ir-6", "related"),
    ("CSF.2.0", "RS.MI-01", "SP.800-53.r5", "ir-4", "related"),
    ("CSF.2.0", "RS.MI-02", "SP.800-53.r5", "ir-4", "related"),
    # Recover
    ("CSF.2.0", "RC.RP-01", "SP.800-53.r5", "cp-10", "related"),
    ("CSF.2.0", "RC.RP-02", "SP.800-53.r5", "cp-10", "related"),
    ("CSF.2.0", "RC.RP-03", "SP.800-53.r5", "cp-9", "related"),
    ("CSF.2.0", "RC.RP-05", "SP.800-53.r5", "cp-10", "related"),
    ("CSF.2.0", "RC.CO-03", "SP.800-53.r5", "ir-6", "related"),
    ("CSF.2.0", "RC.CO-04", "SP.800-53.r5", "ir-6", "related"),
]


# ---------------------------------------------------------------------------
# Public API
# ---------------------------------------------------------------------------


def scrape_mappings(db: sqlite3.Connection) -> int:
    """Download crosswalk XLSX and populate the ``mappings`` table.

    Falls back to hardcoded mappings if the download fails.
    Returns the number of rows inserted.
    """
    rows: list[dict[str, str]] | None = None

    try:
        log.info("Downloading CSF/PF to SP 800-53 crosswalk XLSX ...")
        client = httpx.Client(timeout=_TIMEOUT, follow_redirects=True)
        resp = client.get(_MAPPINGS_URL)
        resp.raise_for_status()
        client.close()

        rows = _parse_mappings_xlsx(resp.content)
        if rows:
            log.info("Parsed %d mappings from crosswalk XLSX", len(rows))
        else:
            log.warning("XLSX parsed but no mappings extracted; falling back")
    except Exception:
        log.warning("Crosswalk download failed; using hardcoded mappings", exc_info=True)

    if rows is None:
        rows = [
            {
                "source_framework": m[0],
                "source_id": m[1],
                "target_framework": m[2],
                "target_id": m[3],
                "relationship": m[4],
            }
            for m in _FALLBACK_MAPPINGS
        ]
        log.info("Using hardcoded mappings: %d entries", len(rows))

    db.execute("DELETE FROM mappings")
    db.executemany(
        """
        INSERT INTO mappings (source_framework, source_id, target_framework,
                              target_id, relationship)
        VALUES (:source_framework, :source_id, :target_framework,
                :target_id, :relationship)
        """,
        rows,
    )
    db.commit()
    log.info("Inserted %d cross-framework mappings", len(rows))
    return len(rows)


# ---------------------------------------------------------------------------
# Standalone test harness
# ---------------------------------------------------------------------------

if __name__ == "__main__":
    logging.basicConfig(level=logging.INFO, format="%(levelname)s: %(message)s")
    conn = sqlite3.connect(":memory:")
    conn.executescript(CREATE_TABLE_SQL)
    n = scrape_mappings(conn)
    print(f"Inserted {n} mappings")

    cur = conn.execute(
        "SELECT source_framework, COUNT(*) FROM mappings GROUP BY source_framework"
    )
    for row in cur:
        print(f"  {row[0]}: {row[1]} mappings")

    cur = conn.execute(
        "SELECT source_id, target_id FROM mappings LIMIT 10"
    )
    for row in cur:
        print(f"  {row[0]} -> {row[1]}")
    conn.close()
