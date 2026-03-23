"""Scraper for NIST publication metadata from CSRC.

Two-step process:
1. Bulk XLSX download from CSRC to get all publication metadata.
2. Optional detail page scraping to enrich with abstract, PDF URL, authors, etc.

The ``--quick`` mode skips step 2, which is useful for testing since step 2
must scrape 2000+ individual pages at 1 req/s.
"""

from __future__ import annotations

import io
import logging
import re
import sqlite3
import time
from typing import Any

import httpx

log = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# Schema
# ---------------------------------------------------------------------------

CREATE_TABLE_SQL = """
CREATE TABLE IF NOT EXISTS publications (
    id              TEXT PRIMARY KEY,
    series          TEXT NOT NULL,
    number          TEXT NOT NULL,
    revision        TEXT,
    title           TEXT NOT NULL,
    abstract        TEXT,
    status          TEXT,
    pub_type        TEXT,
    pub_date        TEXT,
    doi             TEXT,
    pdf_url         TEXT,
    detail_url      TEXT,
    authors         TEXT,          -- comma-separated
    topics          TEXT,          -- comma-separated
    supersedes      TEXT,          -- publication ID
    superseded_by   TEXT,          -- publication ID
    is_latest       INTEGER NOT NULL DEFAULT 1,
    related_pubs    TEXT           -- comma-separated publication IDs
);
CREATE INDEX IF NOT EXISTS idx_pub_series ON publications(series);
CREATE INDEX IF NOT EXISTS idx_pub_status ON publications(status);
CREATE INDEX IF NOT EXISTS idx_pub_date ON publications(pub_date);
CREATE INDEX IF NOT EXISTS idx_pub_latest ON publications(is_latest);

CREATE TABLE IF NOT EXISTS supplemental_materials (
    id          INTEGER PRIMARY KEY AUTOINCREMENT,
    pub_id      TEXT NOT NULL REFERENCES publications(id),
    title       TEXT NOT NULL,
    url         TEXT,
    format      TEXT,
    description TEXT
);
CREATE INDEX IF NOT EXISTS idx_suppl_pub ON supplemental_materials(pub_id);
"""

# ---------------------------------------------------------------------------
# Data source
# ---------------------------------------------------------------------------

_XLSX_URL = (
    "https://csrc.nist.gov/files/pubs/shared/docs/"
    "NIST-Cybersecurity-Publications.xlsx"
)
_DETAIL_BASE = "https://csrc.nist.gov/pubs"
_TIMEOUT = 120
_RATE_LIMIT = 1.0  # seconds between detail page requests


# ---------------------------------------------------------------------------
# XLSX parser
# ---------------------------------------------------------------------------

# Known series names in the XLSX mapped to URL slug / DB prefix
_SERIES_MAP: dict[str, str] = {
    "SP": "sp",
    "SP 800": "sp",
    "SP 1800": "sp",
    "SP 500": "sp",
    "FIPS": "fips",
    "IR": "ir",
    "NISTIR": "ir",
    "CSWP": "cswp",
    "AI": "ai",
    "White Paper": "wp",
}


def _normalize_series(raw: str) -> str:
    """Normalize a series string from the XLSX to a short prefix."""
    raw = raw.strip()
    if raw.startswith("SP 1800"):
        return "SP 1800"
    if raw.startswith("SP 800"):
        return "SP 800"
    if raw.startswith("SP 500"):
        return "SP 500"
    if raw.startswith("SP"):
        return "SP"
    for key in _SERIES_MAP:
        if raw.upper().startswith(key.upper()):
            return key
    return raw


def _make_pub_id(series: str, number: str, revision: str | None) -> str:
    """Construct a publication ID like SP.800-53 or SP.800-53.r5."""
    # Clean up number: strip leading/trailing whitespace
    number = number.strip().replace(" ", "-")
    base = f"{series.replace(' ', '-')}.{number}"
    if revision:
        rev = revision.strip()
        if rev:
            base = f"{base}.{rev}"
    return base


def _parse_xlsx(data: bytes) -> list[dict[str, Any]]:
    """Parse the NIST publications XLSX and return row dicts."""
    from openpyxl import load_workbook

    wb = load_workbook(filename=io.BytesIO(data), read_only=True, data_only=True)
    rows: list[dict[str, Any]] = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]

        # Try to find header row
        header_map: dict[str, int] = {}
        data_start = 0

        all_rows: list[list[str]] = []
        for row in ws.iter_rows(values_only=True):
            str_row = [str(c).strip() if c else "" for c in row]
            all_rows.append(str_row)

        # Scan for header row -- look for cells with keywords like
        # "Series", "Number", "Title", "Status", "Date"
        for i, row_cells in enumerate(all_rows[:10]):
            lower_cells = [c.lower() for c in row_cells]
            # A header row should have at least "title" and one of "series"/"number"
            has_title = any("title" in c for c in lower_cells)
            has_series = any("series" in c for c in lower_cells)
            has_number = any("number" in c or c == "#" or c == "no." for c in lower_cells)
            if has_title and (has_series or has_number):
                for j, cell in enumerate(lower_cells):
                    if "series" in cell:
                        header_map["series"] = j
                    elif cell in ("#", "no.", "no") or "number" in cell:
                        header_map["number"] = j
                    elif cell == "title" or cell == "pub title":
                        header_map["title"] = j
                    elif "rev" in cell:
                        header_map["revision"] = j
                    elif "status" in cell:
                        header_map["status"] = j
                    elif "date" in cell or "released" in cell:
                        header_map["pub_date"] = j
                    elif "type" in cell:
                        header_map["pub_type"] = j
                    elif "doi" in cell:
                        header_map["doi"] = j
                data_start = i + 1
                break

        if not header_map or "title" not in header_map:
            continue

        log.debug(
            "Sheet %s: headers at row %d: %s",
            sheet_name, data_start - 1, header_map,
        )

        for row_cells in all_rows[data_start:]:
            def _get(key: str) -> str:
                idx = header_map.get(key)
                if idx is not None and idx < len(row_cells):
                    return row_cells[idx]
                return ""

            title = _get("title")
            if not title:
                continue

            series_raw = _get("series") or sheet_name
            series = _normalize_series(series_raw)
            number = _get("number")
            revision = _get("revision")
            status = _get("status")
            pub_date = _get("pub_date")
            pub_type = _get("pub_type")
            doi = _get("doi")

            if not number:
                continue

            pub_id = _make_pub_id(series, number, revision)

            rows.append({
                "id": pub_id,
                "series": series,
                "number": number,
                "revision": revision or None,
                "title": title,
                "abstract": None,
                "status": status or None,
                "pub_type": pub_type or None,
                "pub_date": pub_date or None,
                "doi": doi or None,
                "pdf_url": None,
                "detail_url": None,
                "authors": None,
                "topics": None,
                "supersedes": None,
                "superseded_by": None,
                "is_latest": 1,
                "related_pubs": None,
            })

    wb.close()
    return rows


# ---------------------------------------------------------------------------
# Detail page scraper
# ---------------------------------------------------------------------------


def _build_detail_url(series: str, number: str, revision: str | None) -> str:
    """Build the CSRC detail page URL for a publication.

    Real NIST CSRC URLs look like:
        pubs/sp/800/53/r5/upd1/final
        pubs/fips/140-3/final
        pubs/ir/8011/vol-1/final
        pubs/cswp/29/final
    """
    series_slug = _SERIES_MAP.get(series, series.lower().replace(" ", ""))

    # The number field from the XLSX may contain the revision, e.g.
    # "800-53 Rev. 5" or "800-53A Rev. 5".  Split number from revision.
    num_part = number.strip()
    rev_part = revision.strip() if revision else ""

    # Extract revision from number if embedded (e.g., "800-53 Rev. 5")
    rev_match = re.search(r"\s+Rev\.?\s*(\S+)", num_part, re.I)
    if rev_match:
        rev_part = rev_match.group(1)
        num_part = num_part[: rev_match.start()]

    # For SP series: "800-53" needs to become "800/53", "1800-35" -> "1800/35"
    # Split on first hyphen only for the sub-series number
    if series_slug == "sp":
        dash = num_part.find("-")
        if dash > 0:
            prefix = num_part[:dash]      # "800"
            suffix = num_part[dash + 1:]  # "53A"
            # suffix might have more hyphens for volumes: "140A" stays "140A"
            num_path = f"{prefix}/{suffix}"
        else:
            num_path = num_part
    else:
        num_path = num_part

    # Clean up: replace spaces with hyphens, lowercase volume refs
    num_path = num_path.replace(" ", "-").replace("Vol.", "vol").replace("vol-", "vol-")

    url = f"{_DETAIL_BASE}/{series_slug}/{num_path}"

    # Add revision if present: "5" -> "r5", "1" -> "r1"
    if rev_part:
        rev_clean = rev_part.strip().rstrip(".")
        url = f"{url}/r{rev_clean}"

    # NIST CSRC requires a status suffix like /final or /draft
    url = f"{url}/final"

    return url


def _scrape_detail_page(
    client: httpx.Client,
    pub: dict[str, Any],
) -> tuple[dict[str, Any], list[dict[str, Any]]]:
    """Scrape a single publication detail page for enrichment.

    Returns (updated pub dict, list of supplemental material dicts).
    """
    from bs4 import BeautifulSoup

    supplementals: list[dict[str, Any]] = []
    detail_url = _build_detail_url(
        pub["series"], pub["number"], pub.get("revision"),
    )
    pub["detail_url"] = detail_url

    try:
        resp = client.get(detail_url)
        if resp.status_code == 404:
            # Try alternative status suffixes
            for alt_status in ["draft", "ipd"]:
                alt_url = detail_url.rsplit("/", 1)[0] + f"/{alt_status}"
                resp = client.get(alt_url)
                if resp.status_code == 200:
                    detail_url = alt_url
                    pub["detail_url"] = detail_url
                    break
            else:
                # Also try without the /final suffix entirely
                base_url = detail_url.rsplit("/", 1)[0]
                resp = client.get(base_url)
                if resp.status_code != 200:
                    log.debug("Detail page not found: %s", detail_url)
                    return pub, supplementals
                detail_url = base_url
                pub["detail_url"] = detail_url
        resp.raise_for_status()
    except httpx.HTTPError as exc:
        log.debug("Failed to fetch detail page %s: %s", detail_url, exc)
        return pub, supplementals

    soup = BeautifulSoup(resp.text, "lxml")

    # Abstract -- often in a <div class="pub-abstract"> or similar
    abstract_el = soup.find(
        "div", class_=re.compile(r"abstract|pub-abstract", re.I)
    )
    if abstract_el:
        pub["abstract"] = abstract_el.get_text(separator=" ", strip=True)

    # Authors
    authors_el = soup.find("div", class_=re.compile(r"author", re.I))
    if authors_el:
        # Authors might be individual <a> links or a text list
        author_links = authors_el.find_all("a")
        if author_links:
            pub["authors"] = ", ".join(
                a.get_text(strip=True) for a in author_links
            )
        else:
            pub["authors"] = authors_el.get_text(separator=", ", strip=True)

    # PDF URL -- look for links to nvlpubs.nist.gov
    for link in soup.find_all("a", href=True):
        href = link["href"]
        if "nvlpubs.nist.gov" in href and href.endswith(".pdf"):
            pub["pdf_url"] = href
            break
        # Also check doi.org links for DOI
        if "doi.org" in href and not pub.get("doi"):
            pub["doi"] = href

    # Topics -- often in a <div class="pub-topics"> or keyword list
    topics_el = soup.find("div", class_=re.compile(r"topic|keyword", re.I))
    if topics_el:
        topic_links = topics_el.find_all("a")
        if topic_links:
            pub["topics"] = ", ".join(
                a.get_text(strip=True) for a in topic_links
            )

    # Supersedes / Superseded by
    for link in soup.find_all("a", href=True):
        href = link["href"]
        parent = link.find_parent()
        if parent:
            parent_text = parent.get_text(separator=" ", strip=True).lower()
            if "supersedes" in parent_text and "/pubs/" in href:
                pub["supersedes"] = link.get_text(strip=True)
            elif "superseded by" in parent_text and "/pubs/" in href:
                pub["superseded_by"] = link.get_text(strip=True)

    # Supplemental materials -- downloads section
    for dl_section in soup.find_all(
        "div", class_=re.compile(r"suppl|download|companion", re.I)
    ):
        for link in dl_section.find_all("a", href=True):
            href = link["href"]
            if href.startswith("javascript:"):
                continue
            title = link.get_text(strip=True)
            if not title or len(title) < 3:
                continue
            fmt = ""
            if "." in href:
                fmt = href.rsplit(".", 1)[-1].upper()
            supplementals.append({
                "pub_id": pub["id"],
                "title": title,
                "url": href if href.startswith("http") else f"https://csrc.nist.gov{href}",
                "format": fmt,
                "description": None,
            })

    return pub, supplementals


def _compute_is_latest(pubs: list[dict[str, Any]]) -> None:
    """Walk supersedes chains and set is_latest=0 for superseded pubs."""
    # Build a set of all publication IDs that are superseded by something
    superseded_ids: set[str] = set()
    for pub in pubs:
        sup = pub.get("supersedes")
        if sup:
            superseded_ids.add(sup)

    for pub in pubs:
        if pub["id"] in superseded_ids:
            pub["is_latest"] = 0


# ---------------------------------------------------------------------------
# Public API
# ---------------------------------------------------------------------------


def scrape_publications(db: sqlite3.Connection, quick: bool = False) -> int:
    """Download publication metadata. quick=True skips detail page scraping.

    Returns the number of publications inserted.
    """
    client = httpx.Client(
        timeout=_TIMEOUT,
        follow_redirects=True,
        headers={"User-Agent": "nist-mcp-scraper/0.1"},
    )

    # Step 1: Bulk XLSX download
    log.info("Downloading NIST publications XLSX ...")
    try:
        resp = client.get(_XLSX_URL)
        resp.raise_for_status()
    except httpx.HTTPError:
        log.error("Failed to download publications XLSX from %s", _XLSX_URL)
        client.close()
        return 0

    pubs = _parse_xlsx(resp.content)
    log.info("Parsed %d publications from XLSX", len(pubs))

    if not pubs:
        log.warning("No publications parsed from XLSX")
        client.close()
        return 0

    # Step 2: Optional detail page enrichment
    all_supplementals: list[dict[str, Any]] = []
    if not quick:
        log.info("Enriching publications from detail pages (this will take a while) ...")
        for i, pub in enumerate(pubs):
            pub, supplementals = _scrape_detail_page(client, pub)
            all_supplementals.extend(supplementals)
            if (i + 1) % 100 == 0:
                log.info("  Enriched %d / %d publications", i + 1, len(pubs))
            time.sleep(_RATE_LIMIT)
    else:
        log.info("Quick mode: skipping detail page enrichment")
        # Still build detail URLs for reference
        for pub in pubs:
            pub["detail_url"] = _build_detail_url(
                pub["series"], pub["number"], pub.get("revision"),
            )

    client.close()

    # Compute is_latest from supersedes chains
    _compute_is_latest(pubs)

    # Insert into DB
    db.execute("DELETE FROM publications")
    db.execute("DELETE FROM supplemental_materials")

    db.executemany(
        """
        INSERT OR REPLACE INTO publications (
            id, series, number, revision, title, abstract, status,
            pub_type, pub_date, doi, pdf_url, detail_url, authors,
            topics, supersedes, superseded_by, is_latest, related_pubs
        ) VALUES (
            :id, :series, :number, :revision, :title, :abstract, :status,
            :pub_type, :pub_date, :doi, :pdf_url, :detail_url, :authors,
            :topics, :supersedes, :superseded_by, :is_latest, :related_pubs
        )
        """,
        pubs,
    )

    if all_supplementals:
        db.executemany(
            """
            INSERT INTO supplemental_materials (pub_id, title, url, format, description)
            VALUES (:pub_id, :title, :url, :format, :description)
            """,
            all_supplementals,
        )

    db.commit()
    log.info(
        "Inserted %d publications and %d supplemental materials",
        len(pubs), len(all_supplementals),
    )
    return len(pubs)


# ---------------------------------------------------------------------------
# Standalone test harness
# ---------------------------------------------------------------------------

if __name__ == "__main__":
    logging.basicConfig(level=logging.INFO, format="%(levelname)s: %(message)s")
    conn = sqlite3.connect(":memory:")
    conn.executescript(CREATE_TABLE_SQL)
    n = scrape_publications(conn, quick=True)
    print(f"Inserted {n} publications")

    cur = conn.execute(
        "SELECT series, COUNT(*) FROM publications GROUP BY series ORDER BY COUNT(*) DESC"
    )
    for row in cur:
        print(f"  {row[0]}: {row[1]}")

    cur = conn.execute(
        "SELECT id, title FROM publications LIMIT 5"
    )
    for row in cur:
        print(f"  {row[0]}: {row[1][:60]}")
    conn.close()
